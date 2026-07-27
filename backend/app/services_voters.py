"""Voter Excel import and GOTV batch persistence helpers."""

from __future__ import annotations

import hashlib
import logging
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.intelligence.gotv import GOTVPredictor, GOTVProfile, gotv_battleplan
from app.models import Voter

logger = logging.getLogger(__name__)

COLUMN_ALIASES = {
    "first_name": {"first_name", "firstname", "first", "שם פרטי", "שם_פרטי"},
    "last_name": {"last_name", "lastname", "last", "שם משפחה", "שם_משפחה", "family"},
    "city": {"city", "עיר"},
    "neighborhood": {"neighborhood", "שכונה", "area"},
    "phone": {"phone", "mobile", "טלפון", "נייד"},
    "email": {"email", "mail", "אימייל"},
    "support_score": {"support_score", "support", "score", "תמיכה"},
    "turnout_history": {"turnout_history", "turnout", "הצבעה"},
    "notes": {"notes", "הערות", "note"},
}


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = _norm_header(header)
        for field, aliases in COLUMN_ALIASES.items():
            if key in aliases or key == field:
                mapping[field] = idx
                break
    # Hebrew single-column "שם" → split later
    if "first_name" not in mapping:
        for idx, header in enumerate(headers):
            if _norm_header(header) in {"name", "שם", "full_name", "fullname_name"}:
                mapping["full_name"] = idx
                break
    return mapping


def _synthetic_national_id(first: str, last: str, phone: str | None) -> str:
    raw = f"{first}|{last}|{phone or ''}".encode()
    return hashlib.sha1(raw).hexdigest()[:16]


def parse_excel_voters(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    mapping = _map_headers(headers)
    voters: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict[str, Any] = {}
        if "full_name" in mapping and "first_name" not in mapping:
            full = str(row[mapping["full_name"]] or "").strip()
            parts = full.split(None, 1)
            record["first_name"] = parts[0] if parts else "Unknown"
            record["last_name"] = parts[1] if len(parts) > 1 else "-"
        else:
            record["first_name"] = str(row[mapping["first_name"]] if "first_name" in mapping else "").strip() or "Unknown"
            record["last_name"] = str(row[mapping["last_name"]] if "last_name" in mapping else "").strip() or "-"
        for field in ("city", "neighborhood", "phone", "email", "notes"):
            if field in mapping:
                value = row[mapping[field]]
                record[field] = None if value is None else str(value).strip()
        for field in ("support_score", "turnout_history"):
            if field in mapping and row[mapping[field]] is not None:
                try:
                    record[field] = float(row[mapping[field]])
                except (TypeError, ValueError):
                    record[field] = None
        voters.append(record)
    return voters


def import_voters(db: Session, records: list[dict[str, Any]]) -> dict[str, int]:
    imported = 0
    duplicates = 0
    for record in records:
        first = record["first_name"]
        last = record["last_name"]
        existing = (
            db.query(Voter)
            .filter(Voter.first_name == first, Voter.last_name == last)
            .one_or_none()
        )
        if existing:
            duplicates += 1
            continue
        voter = Voter(
            national_id=_synthetic_national_id(first, last, record.get("phone")),
            first_name=first,
            last_name=last,
            city=record.get("city"),
            neighborhood=record.get("neighborhood"),
            phone=record.get("phone"),
            email=record.get("email"),
            notes=record.get("notes"),
            support_score=record.get("support_score"),
            turnout_history=record.get("turnout_history"),
            turnout_score=record.get("turnout_history"),
        )
        db.add(voter)
        imported += 1
    db.commit()
    return {"imported": imported, "duplicates": duplicates, "total": imported + duplicates}


def apply_gotv_to_voter(voter: Voter, profile: GOTVProfile) -> None:
    voter.gotv_category = profile.category.value
    voter.gotv_priority = profile.priority_score
    voter.gotv_channel = profile.optimal_channel.value
    voter.gotv_frequency = profile.contact_frequency
    voter.gotv_message = profile.messaging_frame
    voter.updated_at = datetime.now(UTC)


def classify_db_voters(db: Session, predictor: GOTVPredictor | None = None) -> dict[str, Any]:
    predictor = predictor or GOTVPredictor()
    voters = db.query(Voter).all()
    payload = [
        {
            "name": f"{v.first_name} {v.last_name}".strip(),
            "support_score": v.support_score if v.support_score is not None else 0.5,
            "turnout_history": v.turnout_history
            if v.turnout_history is not None
            else (v.turnout_score if v.turnout_score is not None else 0.55),
        }
        for v in voters
    ]
    profiles = predictor.classify_batch(payload)
    by_name = {p.name: p for p in profiles}
    for voter in voters:
        name = f"{voter.first_name} {voter.last_name}".strip()
        profile = by_name.get(name)
        if profile:
            apply_gotv_to_voter(voter, profile)
    db.commit()
    battle = gotv_battleplan(profiles)
    categories = {k.lower(): len(v) for k, v in (battle.get("segments") or {}).items()}
    return {
        "classified": len(profiles),
        "categories": {
            "safe": categories.get("safe", 0),
            "leaning": categories.get("leaning", 0),
            "swing": categories.get("swing", 0),
            "at_risk": categories.get("at_risk", 0),
            "lost": categories.get("lost", 0),
        },
        "battle_plan": {
            "field_ops": battle.get("resource_allocation", {}),
            "channels": battle.get("resource_allocation", {}),
            "top_swing": [
                item for item in battle.get("top_priority", []) if item.get("category") == "SWING"
            ][:20],
            "top_priority": battle.get("top_priority", []),
            "segments": battle.get("segments", {}),
            "summary": battle.get("summary", {}),
        },
        "voters": [
            {
                "name": p.name,
                "category": p.category.value,
                "priority_score": p.priority_score,
                "optimal_channel": p.optimal_channel.value,
                "contact_frequency": p.contact_frequency,
                "messaging_frame": p.messaging_frame,
                "turnout_probability": p.turnout_probability,
                "persuasion_score": p.persuasion_score,
            }
            for p in sorted(profiles, key=lambda x: x.priority_score, reverse=True)
        ],
    }
